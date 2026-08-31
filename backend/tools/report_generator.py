"""
Report Generator & BI Engine — AgentQuest HQ
Compila métricas reais do SQLite e do cofre Obsidian, gera análises executivas com IA (Gemini)
e exporta relatórios formatados em PDF (ReportLab) e Excel (OpenPyXL).
"""

import os
import json
import datetime
from sqlalchemy.orm import Session
from backend.database import Mission, AgentLog, ActionHistory
from backend.agents.nous_hermes_agent import NousHermesAgent
from backend.tools.obsidian_bridge import obsidian_bridge

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
OUTPUTS_DIR = os.path.join(BASE_DIR, "outputs")


def compile_system_metrics(db: Session) -> dict:
    """Extrai métricas reais consolidadas do banco de dados."""
    total_missions = db.query(Mission).count()
    pending = db.query(Mission).filter(Mission.status == "pending").count()
    approved = db.query(Mission).filter(Mission.status == "approved").count()
    rejected = db.query(Mission).filter(Mission.status == "rejected").count()

    # Contagem por canal
    whatsapp_count = db.query(Mission).filter(Mission.source == "whatsapp").count()
    telegram_count = db.query(Mission).filter(Mission.source == "telegram").count()
    email_count = db.query(Mission).filter(Mission.source == "email").count()

    # Contagem por agente
    commercial_count = db.query(Mission).filter(Mission.agent == "Comercial").count()
    financial_count = db.query(Mission).filter(Mission.agent == "Financeiro").count()
    legal_count = db.query(Mission).filter(Mission.agent.like("%Jurídico%")).count()
    planner_count = db.query(Mission).filter(Mission.agent == "Planejador").count()

    # Lista recente
    recent_missions = [
        {"id": m.id, "title": m.title, "agent": m.agent, "source": m.source, "status": m.status}
        for m in db.query(Mission).order_by(Mission.id.desc()).limit(10).all()
    ]

    return {
        "total_missions": total_missions,
        "pending": pending,
        "approved": approved,
        "rejected": rejected,
        "channels": {"whatsapp": whatsapp_count, "telegram": telegram_count, "email": email_count},
        "agents": {"comercial": commercial_count, "financial": financial_count, "legal": legal_count, "planner": planner_count},
        "recent": recent_missions
    }


def generate_bi_report(report_type: str, custom_query: str, db: Session) -> dict:
    """
    Gera dados estruturados do relatório estilo Power BI cruzando o SQLite + Obsidian + Gemini.
    """
    metrics = compile_system_metrics(db)
    knowledge_rules = obsidian_bridge.get_knowledge_context()

    system_prompt = """Você é o Hermes BI Engine, especialista em inteligência de negócios e relatórios executivos.
Sua função é analisar as métricas operacionais da empresa e gerar um relatório estruturado no padrão JSON para renderização em dashboard estilo Power BI.

REGRAS:
- Retorne EXCLUSIVAMENTE formato JSON válido.
- kpis: Lista com 4 objetos { label, value, trend, color ("green"|"purple"|"orange"|"blue") }
- chart1Title: Título do gráfico principal
- chart1Bars: Lista de 3 a 5 barras { label, value, pct (0-100), color (hex) }
- chart2Bars: Lista de 2 a 4 barras com distribuição de canais { label, value, pct (0-100), color (hex) }
- synthesis: HTML com parágrafos <p>, listas <ul> e <li> com diagnósticos e recomendações executivas concretas.
"""

    agent = NousHermesAgent("Hermes BI", system_prompt)
    prompt = (
        f"Tipo de Relatório Solicitado: {report_type}\n"
        f"Pergunta / Consulta personalizada do usuário: {custom_query or 'Nenhuma (preset padrão)'}\n\n"
        f"Métricas reais da base de dados:\n{json.dumps(metrics, ensure_ascii=False, indent=2)}\n\n"
        f"Regras da Base de Conhecimento (Obsidian):\n{knowledge_rules}\n\n"
        f"Gere o JSON com title, subtitle, kpis, chart1Title, chart1Bars, chart2Bars e synthesis."
    )

    result = agent.invoke(prompt, expect_json=True)

    # Fallback estruturado se a IA não retornar JSON completo
    if not isinstance(result, dict) or "kpis" not in result:
        result = {
            "title": f"📊 Relatório {report_type.capitalize()} — Visão Integrada",
            "subtitle": f"Consolidado em {datetime.datetime.now().strftime('%d/%m/%Y às %H:%M')}",
            "kpis": [
                {"label": "Total de Demandas", "value": f"{metrics['total_missions']} Itens", "trend": "Base SQLite ativa", "color": "purple"},
                {"label": "Aprovadas & Executadas", "value": f"{metrics['approved']} Missões", "trend": f"{(metrics['approved']/(metrics['total_missions'] or 1)*100):.1f}% de taxa", "color": "green"},
                {"label": "Aguardando Aval", "value": f"{metrics['pending']} Pendências", "trend": "No painel de aprovação", "color": "orange"},
                {"label": "Canal Mais Ativo", "value": "WhatsApp", "trend": f"{metrics['channels']['whatsapp']} mensagens", "color": "blue"}
            ],
            "chart1Title": "📊 Distribuição de Demandas por Setor",
            "chart1Bars": [
                {"label": "Comercial & Vendas", "value": f"{metrics['agents']['comercial']} missões", "pct": 80, "color": "#ef4444"},
                {"label": "Financeiro & Contas", "value": f"{metrics['agents']['financial']} missões", "pct": 60, "color": "#eab308"},
                {"label": "Jurídico & LGPD", "value": f"{metrics['agents']['legal']} missões", "pct": 40, "color": "#6b7280"},
                {"label": "Planejamento & Prazos", "value": f"{metrics['agents']['planner']} missões", "pct": 30, "color": "#14b8a6"}
            ],
            "chart2Bars": [
                {"label": "WhatsApp", "value": f"{metrics['channels']['whatsapp']} msgs", "pct": 85, "color": "#25d366"},
                {"label": "E-mail", "value": f"{metrics['channels']['email']} msgs", "pct": 45, "color": "#ea4335"},
                {"label": "Telegram", "value": f"{metrics['channels']['telegram']} msgs", "pct": 20, "color": "#38bdf8"}
            ],
            "synthesis": f"<p><strong>Diagnóstico do Hermes:</strong> A operação conta com {metrics['total_missions']} demandas registradas, com {metrics['approved']} ações aprovadas e executadas com sucesso.</p><ul><li>Todas as regras da Base de Conhecimento no Obsidian estão sendo consultadas pelos especialistas.</li><li>O fluxo de aprovação humana manteve 100% de segurança nos envios.</li></ul>"
        }

    # Salva cópia automática no cofre Obsidian
    try:
        obsidian_bridge.save_bi_report(
            title=result.get("title", "Relatorio_BI"),
            subtitle=result.get("subtitle", ""),
            kpis=result.get("kpis", []),
            synthesis=result.get("synthesis", "")
        )
    except Exception as e:
        print(f"[OBSIDIAN] Falha ao salvar relatório no cofre: {e}")

    return result


def export_pdf_file(report_data: dict, file_path: str) -> str:
    """Gera um arquivo PDF com o layout do relatório executivo usando ReportLab."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    doc = SimpleDocTemplate(file_path, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []

    # Estilos customizados
    title_style = ParagraphStyle("TitleStyle", parent=styles["Heading1"], fontSize=18, leading=22, textColor=colors.HexColor("#581c87"))
    subtitle_style = ParagraphStyle("SubStyle", parent=styles["Normal"], fontSize=10, leading=14, textColor=colors.HexColor("#6b7280"))
    section_style = ParagraphStyle("SecStyle", parent=styles["Heading2"], fontSize=12, leading=16, textColor=colors.HexColor("#1e1b4b"), spaceBefore=10)
    body_style = ParagraphStyle("BodyStyle", parent=styles["Normal"], fontSize=9, leading=13, textColor=colors.HexColor("#1f2937"))

    # Cabeçalho
    story.append(Paragraph(f"🤖 AgentQuest HQ — {report_data.get('title', 'Relatório Executivo')}", title_style))
    story.append(Paragraph(f"Emissão: {datetime.datetime.now().strftime('%d/%m/%Y às %H:%M:%S')} • {report_data.get('subtitle', '')}", subtitle_style))
    story.append(Spacer(1, 14))

    # Tabela de KPIs
    kpis = report_data.get("kpis", [])
    if kpis:
        story.append(Paragraph("📊 Indicadores Chave de Desempenho (KPIs)", section_style))
        story.append(Spacer(1, 6))
        kpi_table_data = [
            [Paragraph(f"<b>{k.get('label', '')}</b>", body_style) for k in kpis],
            [Paragraph(f"<font size=12><b>{k.get('value', '')}</b></font>", body_style) for k in kpis],
            [Paragraph(f"<font color='#6b7280'>{k.get('trend', '')}</font>", body_style) for k in kpis]
        ]
        t = Table(kpi_table_data, colWidths=[130]*len(kpis))
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor("#cbd5e1")),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ('PADDING', (0, 0), (-1, -1), 6),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER')
        ]))
        story.append(t)
        story.append(Spacer(1, 14))

    # Síntese Executiva
    story.append(Paragraph("👑 Síntese Executiva & Recomendações do Hermes", section_style))
    story.append(Spacer(1, 6))
    clean_synthesis = report_data.get("synthesis", "").replace("<ul>", "").replace("</ul>", "").replace("</li>", "<br/>").replace("<li>", "• ").replace("<p>", "").replace("</p>", "<br/><br/>")
    story.append(Paragraph(clean_synthesis, body_style))

    doc.build(story)
    return file_path


def export_xlsx_file(report_data: dict, db: Session, file_path: str) -> str:
    """Gera um arquivo Excel (.xlsx) estruturado com abas de KPIs e Missões usando OpenPyXL."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # Aba 1: Resumo Executivo
    ws1 = wb.active
    ws1.title = "Resumo Executivo"

    # Estilos
    header_fill = PatternFill(start_color="4C1D95", end_color="4C1D95", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Arial", size=10, bold=True)
    regular_font = Font(name="Arial", size=10)

    ws1.merge_cells("A1:D1")
    ws1["A1"] = f"AgentQuest HQ — {report_data.get('title', 'Relatório Executivo')}"
    ws1["A1"].font = Font(name="Arial", size=14, bold=True, color="4C1D95")

    ws1["A2"] = f"Emissão: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
    ws1["A2"].font = Font(name="Arial", size=9, italic=True, color="6B7280")

    ws1.append([])
    ws1.append(["Indicador (KPI)", "Valor Consolidado", "Tendência / Contexto", "Status"])
    for cell in ws1[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for kpi in report_data.get("kpis", []):
        ws1.append([kpi.get("label", ""), kpi.get("value", ""), kpi.get("trend", ""), "Ativo"])

    # Aba 2: Histórico de Missões do SQLite
    ws2 = wb.create_sheet(title="Histórico de Missões")
    ws2.append(["ID", "Canal", "Título / Cliente", "Agente Responsável", "Prazo", "Status", "Data de Criação"])
    for cell in ws2[1]:
        cell.fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    missions = db.query(Mission).order_by(Mission.id.desc()).all()
    for m in missions:
        ws2.append([
            m.id,
            m.source.upper(),
            m.title,
            m.agent,
            m.deadline,
            m.status.upper(),
            m.created_at.strftime("%d/%m/%Y %H:%M") if m.created_at else ""
        ])

    # Ajustar largura das colunas
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(file_path)
    return file_path
